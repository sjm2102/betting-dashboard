# generate_excel.py — builds edge_bets.xlsx from bets.json
# Runs after resolve_bets.js in GitHub Actions
import json, sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

BETS_FILE   = 'bets.json'
EXCEL_FILE  = 'edge_bets.xlsx'

# ── Colors ────────────────────────────────────────────────────
BG_HEADER  = 'FF0A0A0A'  # near black
FG_HEADER  = 'FF00E676'  # green
BG_TRUE    = 'FF0D2B1A'  # dark green
FG_TRUE    = 'FF00E676'
BG_FALSE   = 'FF2B0D0D'  # dark red
FG_FALSE   = 'FFFF3D57'
BG_PENDING = 'FF1A1A1A'
FG_PENDING = 'FFCCAA00'
BG_ROW_ALT = 'FF141414'
BG_ROW     = 'FF111111'
FG_TEXT    = 'FFE8E8E8'
FG_DIM     = 'FF888888'

def cell_fill(color): return PatternFill('solid', start_color=color, end_color=color)
def cell_font(color, bold=False, size=9): return Font(name='Arial', color=color, bold=bold, size=size)
def thin_border():
    s = Side(border_style='thin', color='FF2A2A2A')
    return Border(left=s, right=s, top=s, bottom=s)
def center(): return Alignment(horizontal='center', vertical='center', wrap_text=False)

COLUMNS = [
    ('Date',        12, 'date'),
    ('Player',      22, 'player'),
    ('Market',       7, 'market'),
    ('Side',         7, 'side'),
    ('Line',         7, 'line'),
    ('Actual',       8, 'actualStat'),
    ('Result',       8, 'result'),
    ('Edge Score',  11, 'edgeScore'),
    ('EV',           8, 'ev'),
    ('DVP',          6, 'dvpGrade'),
    ('Season Avg',  11, 'seasonAvg'),
    ('MPG',          6, 'mpg'),
    ('USG%',         7, 'usg'),
    ('Odds',         7, 'odds'),
    ('Book',        12, 'book'),
    ('Game',        22, 'game'),
]

def build_excel(bets):
    wb = Workbook()

    # ── MAIN SHEET: All Bets ─────────────────────────────────
    ws = wb.active
    ws.title = 'All Bets'
    ws.freeze_panes = 'A2'
    ws.sheet_view.showGridLines = False

    # Header row
    ws.row_dimensions[1].height = 22
    for col_idx, (label, width, _) in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=col_idx, value=label)
        c.font      = cell_font(FG_HEADER, bold=True, size=9)
        c.fill      = cell_fill(BG_HEADER)
        c.alignment = center()
        c.border    = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows
    for row_idx, bet in enumerate(reversed(bets), 2):
        ws.row_dimensions[row_idx].height = 17
        bg = BG_ROW_ALT if row_idx % 2 == 0 else BG_ROW
        result = bet.get('result', 'PENDING')

        for col_idx, (_, _, key) in enumerate(COLUMNS, 1):
            val = bet.get(key, '')
            if val is None: val = ''
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.fill      = cell_fill(bg)
            c.alignment = center()
            c.border    = thin_border()

            # Result column coloring
            if key == 'result':
                if result == 'TRUE':
                    c.fill = cell_fill(BG_TRUE)
                    c.font = cell_font(FG_TRUE, bold=True)
                elif result == 'FALSE':
                    c.fill = cell_fill(BG_FALSE)
                    c.font = cell_font(FG_FALSE, bold=True)
                else:
                    c.fill = cell_fill(BG_PENDING)
                    c.font = cell_font(FG_PENDING, bold=True)
            elif key == 'player':
                c.font = cell_font(FG_TEXT, bold=True)
                c.alignment = Alignment(horizontal='left', vertical='center')
            elif key == 'edgeScore':
                c.font = cell_font(FG_HEADER, bold=True)
            else:
                c.font = cell_font(FG_TEXT)

    # ── SUMMARY SHEET ────────────────────────────────────────
    ws2 = wb.create_sheet('Summary')
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions['A'].width = 24
    ws2.column_dimensions['B'].width = 16

    settled = [b for b in bets if b.get('result') in ('TRUE','FALSE')]
    true_count  = sum(1 for b in settled if b['result'] == 'TRUE')
    false_count = sum(1 for b in settled if b['result'] == 'FALSE')
    pending_count = sum(1 for b in bets if b.get('result') == 'PENDING')
    hit_rate = (true_count / len(settled) * 100) if settled else 0

    by_market = {}
    for b in settled:
        m = b.get('market','?')
        if m not in by_market: by_market[m] = {'true':0,'false':0}
        by_market[m]['true'  if b['result']=='TRUE' else 'false'] += 1

    rows = [
        ('THE EDGE — Bet Tracker', None),
        (f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}', None),
        ('', None),
        ('OVERALL', None),
        ('Total Bets Logged', len(bets)),
        ('Settled',           len(settled)),
        ('Pending',           pending_count),
        ('TRUE (Hit)',        true_count),
        ('FALSE (Miss)',      false_count),
        ('Hit Rate',          f'{hit_rate:.1f}%'),
        ('', None),
        ('BY MARKET', None),
    ]
    for m, counts in sorted(by_market.items()):
        total = counts['true'] + counts['false']
        rate  = counts['true'] / total * 100 if total else 0
        rows.append((m, f"{counts['true']}W / {counts['false']}L  ({rate:.1f}%)"))

    for r_idx, (label, val) in enumerate(rows, 1):
        ws2.row_dimensions[r_idx].height = 18
        ca = ws2.cell(row=r_idx, column=1, value=label)
        cb = ws2.cell(row=r_idx, column=2, value=val if val is not None else '')
        for c in [ca, cb]:
            c.fill = cell_fill(BG_ROW)
            c.border = thin_border()

        if label in ('THE EDGE — Bet Tracker',):
            ca.font = cell_font(FG_HEADER, bold=True, size=13)
        elif label in ('OVERALL', 'BY MARKET'):
            ca.font = cell_font(FG_HEADER, bold=True)
        elif label == 'Hit Rate':
            cb.font = cell_font(FG_TRUE if hit_rate >= 50 else FG_FALSE, bold=True)
            ca.font = cell_font(FG_TEXT)
        else:
            ca.font = cell_font(FG_DIM if not val else FG_TEXT)
            cb.font = cell_font(FG_TEXT, bold=True)

    wb.save(EXCEL_FILE)
    print(f'✅ Saved {EXCEL_FILE} — {len(bets)} bets ({true_count}T/{false_count}F/{pending_count}P)')

if __name__ == '__main__':
    if not __import__('os').path.exists(BETS_FILE):
        print('No bets.json — nothing to generate.'); sys.exit(0)
    bets = json.load(open(BETS_FILE))
    build_excel(bets)
